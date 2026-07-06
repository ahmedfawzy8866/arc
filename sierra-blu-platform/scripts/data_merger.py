import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────
#  CONFIG
# ──────────────────────────────────────────
SOURCE_FOLDER  = r"I:\Alpha"  # Directory containing files to merge
TEMPLATE_FILE  = r"I:\Beta\17-4-2026.xlsx"  # Template file for columns
OUTPUT_FILE    = r"I:\Beta\SBmerged_output.xlsx"  # Final output file
DEDUP_COLUMN   = "Mobile"  # Deduplication column
TEST_MODE      = True
TEST_ROWS      = 5

OWNERS_RENT_KEYWORDS   = ["مالك", "owner", "rent", "إيجار"]
OWNERS_RESALE_KEYWORDS = ["owner", "resale", "sale", "بيع"]
BROKERS_KEYWORDS       = ["بروكر", "broker", "وسيط"]
TEAM_KEYWORDS          = ["team", "فريق", "team units"]

SHEETS = {
    "Owners-Rent": [],
    "Owners-Resale": [],
    "Brokers": [],
    "Team Units": [],
    "blanc": []
}

def get_template_columns(path):
    try:
        df = pd.read_excel(path, sheet_name=0, nrows=0)
        return list(df.columns)
    except Exception as e:
        print(f"❌ Error reading template: {e}")
        return None

def find_classify_column(df):
    all_kw = OWNERS_RENT_KEYWORDS + OWNERS_RESALE_KEYWORDS + BROKERS_KEYWORDS + TEAM_KEYWORDS
    for col in df.columns:
        for kw in all_kw:
            if kw.lower() in str(col).lower(): return col
    for col in df.columns:
        try:
            sample = df[col].dropna().astype(str)
            for kw in all_kw:
                if sample.str.contains(kw, case=False, na=False).any(): return col
        except: continue
    return None

def classify_row(value):
    if pd.isna(value) or str(value).strip() == "": return "blanc"
    val = str(value).strip().lower()
    for kw in OWNERS_RENT_KEYWORDS:
        if kw.lower() in val: return "Owners-Rent"
    for kw in OWNERS_RESALE_KEYWORDS:
        if kw.lower() in val: return "Owners-Resale"
    for kw in BROKERS_KEYWORDS:
        if kw.lower() in val: return "Brokers"
    for kw in TEAM_KEYWORDS:
        if kw.lower() in val: return "Team Units"
    return "blanc"

def read_all_files(folder, template_cols, test_mode=False):
    sheets_data = {k: [] for k in SHEETS.keys()}
    files = [f for f in os.listdir(folder) if f.lower().endswith(".xlsx")]
    
    if not files: return sheets_data

    for fname in files:
        fpath = os.path.join(folder, fname)
        try: xl = pd.ExcelFile(fpath)
        except: continue

        for sheet in xl.sheet_names:
            try: 
                df = xl.parse(sheet)
                raw_df = df.copy()
            except: continue

            if df.empty: continue
            if test_mode:
                df = df.head(TEST_ROWS)
                raw_df = raw_df.head(TEST_ROWS)

            classify_col = find_classify_column(df)
            for col in template_cols:
                if col not in df.columns: df[col] = None
            df = df[template_cols]

            for i in range(len(df)):
                val = raw_df[classify_col].iloc[i] if classify_col and classify_col in raw_df.columns else ""
                category = classify_row(val)
                sheets_data[category].append(df.iloc[i].to_dict())
    return sheets_data

def to_df_dedup(rows, cols, dedup_col):
    if not rows: return pd.DataFrame(columns=cols)
    df = pd.DataFrame(rows, columns=cols).fillna("N/A")
    if dedup_col in df.columns:
        df = df.drop_duplicates(subset=[dedup_col], keep="first")
    return df

def write_output(path, sheets_data, dedup_col):
    dfs = {name: to_df_dedup(rows, list(rows[0].keys()) if rows else [], dedup_col) for name, rows in sheets_data.items()}
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in dfs.items(): df.to_excel(writer, sheet_name=name, index=False)
    
    wb = load_workbook(path)
    header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    for name in dfs.keys():
        ws = wb[name]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
            ws.column_dimensions[col_letter].width = min(max_len + 3, 50)
    wb.save(path)

def main():
    template_cols = get_template_columns(TEMPLATE_FILE)
    if not template_cols: return
    sheets_data = read_all_files(SOURCE_FOLDER, template_cols, test_mode=TEST_MODE)
    write_output(OUTPUT_FILE, sheets_data, DEDUP_COLUMN)

if __name__ == "__main__":
    main()
